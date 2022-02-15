import pandas as pd 
import datetime as dt 
import time as tm 
import statistics
from statistics import mean,mode,median
pd.options.mode.chained_assignment = None
import numpy as np
import os 
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl import load_workbook as load 
lagbe = ""
def TheGrandSorting(choose):
	dic = {0:'A',1:'B',2:'C',3:'D',4:'E',5:'F',6:'G',7:'H',8:'I',9:'J',10:'K',11:'L',12:'M',13:'N',14:'O',15:'P',16:'Q',17:'R',18:'S',19:'T',20:'U',21:'V',22:'W',23:'X',24:'Y',25:'Z',26:'AA',27:'AB',28:'AC',29:'AD',30:'AE',31:'AF',32:'AG',33:'AH',34:'AI',35:'AJ',36:'AK',37:'AL',38:'AM',39:'AN',40:'AO',41:'AP',42:'AQ',43:'AR',44:'AS',45:'AT',46:'AU',47:'AV',48:'AW',49:'AX',50:'AY',51:'AZ',52:'BA'}
	mon = {1:'Jan',2:'Feb',3:'Mar',4:'April',5:'May',6:"June",7:"July",8:'August',9:'September',10:"October",11:'November',12:'December'}
	path = '/home/serenity/Dropbox/Farah Checks/Pending/Core Questions Queries/'
	file_addr = os.listdir('/home/serenity/Dropbox/Farah Checks/Pending/Core Questions Queries')
	# file_select = [nam for nam in file_addr if nam[-17:] == '(all months).xlsx'] 
	# file_select = [nam for nam in file_addr if nam[-20:] == 'Inconsistencies.xlsx' or nam[-17:] == '(all months).xlsx']
	# file_select = [nam for nam in file_addr if nam[-20:] == 'Inconsistencies.xlsx' or nam[-17:] == 'Data Queries.xlsx' or nam[-17:] == '(all months).xlsx']
	file_select = [nam for nam in file_addr if nam[-20:] == 'Inconsistencies.xlsx' or nam[-12:] == 'Queries.xlsx' or nam[-17:] == '(all months).xlsx' or nam[-9:] == 'Name.xlsx']

	tabularasa = pd.DataFrame()
	dictionary = {'name':tabularasa}
	sanem = pd.read_excel('Sanem MFO Team.xlsx')
	lis = []
	sup = []
	for file in file_select:
		xls = pd.ExcelFile(path+file)
		# workbook = load(path+file, data_only = True)
		for tab in xls.sheet_names:
			df = pd.read_excel(path+file,sheet_name = tab)
			
			for resp in df['respid']:
				# print(resp,tab)
				print(resp,type(resp),file,tab)
				# if type(resp) == 'float':
				# 	print(resp)
				if resp[:3] in sanem[sanem['supervisors'] == 'Sup ' + str(choose)].enumerators.values :
					lis.append(resp)
			df = df[df['respid'].isin(lis)]
			df.insert(0,'Sheet_Name',tab)
			df.insert(0,'File_Name',file)
			if tab[-5:] == 'clean' or df.empty:
				continue
			else:
				try:
					# dictionary.update({ file + " Sheet: " + tab :df[ ( (df['status'] == 'pending') & ( df['answer'].isnull() ) ) |  (df['status'] == 'queried') & ( df['answer_1'].isnull() ) |  (df['status'] == 'queried') & ( df['answer_2'].isnull() ) |  (df['status'] == 'double-queried') & ( df['answer_1'].isnull() )|  (df['status'] == 'triple-queried') & ( df['answer_2'].isnull() ) ] })
					dfnew = df[ ( (df['status'] == 'pending') & ( df['answer'].isnull() ) ) |  (df['status'] == 'queried') & ( df['answer_1'].isnull() ) | (df['status'] == 'double-queried') & ( df['answer_2'].isnull() ) ]
					dfnew.insert(0,'Index', '#') #new
					dfnew.loc[0] = dfnew.columns #brings the column name to the first row of each dataframe
					dfnew.columns = range(len(dfnew.columns)) #place column names as 0 to length of the dataframe for each dataframe, this algorithm makes evry dataframe to make stack onto each other
					
					if len(dfnew) == 1:
						continue
					else:
						dictionary.update({ file + " Sheet: " + tab : dfnew.sort_index() })
					# dictionary.update({ file + " Sheet: " + tab :df[ ( (df['status'] == 'pending') & ( df['answer'].isnull() ) ) |  (df['status'] == 'queried') & ( df['answer_1'].isnull() ) | (df['status'] == 'double-queried') & ( df['answer_2'].isnull() ) ] })
					
				except:
					dfnew = df[ (df['status'] == 'pending') & ( df['answer'].isnull() ) ]
					dfnew.insert(0,'Index', '#' ) #new
					dfnew.loc[0] = dfnew.columns
					dfnew.columns = range(len(dfnew.columns))
					
					if len(dfnew) == 1:
						continue
					else:
						dictionary.update({ file + " Sheet: " + tab : dfnew.sort_index()}) 
				# dictionary.update({ file + " Sheet: " + tab :df [ (df['status'] != 'solved')  ] })

	tabularasa = pd.concat(dictionary.values(),axis = 0, keys=dictionary.keys())
	# for i in tabularasa.index:

	# 	print(i[1])
	cou = 0
	# for co in tabularasa[0][0]:
	# 	print(co)
	# print("Supervisor " + str(choose)+" "+ cou)
	# for som in tabularasa[0]:
	# 	if som == "#":
	# 		cou = cou + 1
	# print("Supervisor " + str(choose)+" "+ str(cou) )
	# 
	tabularasa.to_excel("Supervisor " + str(choose) + " queries" + ".xlsx",encoding = 'utf-08',index= True,header = False)
	workbook = load("Supervisor " + str(choose) + " queries" + ".xlsx")
	ws = workbook["Sheet1"]
	# thin = Side(border_style="thin",color = "B2BEB5")
	thin = Side(border_style="thin",color = "98FB98")
	for i in range(1,100):
		if ws['B'+str(i)].value == 0:
			for z in range(52):
				ws[ dic[z] + str(i) ].font = Font(bold = True)
				ws[ dic[z] + str(i) ].border = Border(bottom = thin)
				# ws[ dic[z] + str(i) ].fill = PatternFill("solid",fgColor = "98FB98")

	for i in range(1,100):
		for z in range(52):
			if ws[ dic[z] + str(i) ].value == 'queried':
				for z in range(52):
					ws[ dic[z] + str(i) ].fill = PatternFill("solid",fgColor = "FFFF00")
	for i in range(1,100):
		for z in range(52):
			if ws[ dic[z] + str(i) ].value == 'double-queried':
				for z in range(52):
					ws[ dic[z] + str(i) ].fill = PatternFill("solid",fgColor = "FFFFC000")

	for i in range(1,100):
		for z in range(52):
			if ws[ dic[z] + str(i) ].value == 'answer':
				ws[ dic[z] + str(i) ].fill = PatternFill("solid",fgColor = "FF00B0F0")
	for i in range(1,100):
		for z in range(52):
			if ws[ dic[z] + str(i) ].value == 'answer_1':
				ws[ dic[z] + str(i) ].fill = PatternFill("solid",fgColor = "FF00B0F0")
	# openpyxl.worksheet.dimensions.ColumnDimension(ws,bestFit = False, width=10)
	for i in range(1,100):
		for z in range(52):
			if ws[ dic[z] + str(i) ].value == 'answer_2':
				ws[ dic[z] + str(i) ].fill = PatternFill("solid",fgColor = "FF00B0F0")

	ws.delete_cols(1)
	dtnow = dt.datetime.fromtimestamp(tm.time())
	ws["B100"].value = "updated on " + str(dtnow.day) + " "+ mon[dtnow.month] + " at " + str(dtnow.hour) + ":" + str(dtnow.minute)
	# def as_text(value):
	#     if value is None:
	#         return ""
	#     return str(value)
	# for column_cells in ws.columns:
	#     length = max(len(as_text(cell.value)) for cell in column_cells)
	#     ws.column_dimensions[column_cells[0].column_letter].width = length
	for z in range(52):
		ws.column_dimensions[dic[z]].width =  mean( len( str(ws[dic[z] + str(i)].value)) for i in range(1,100) ) + mode( len( str(ws[dic[z] + str(i)].value)) for i in range(1,100) ) + median( len( str(ws[dic[z] + str(i)].value)) for i in range(1,100) )
	for i in range(1,100):
		for z in range(52):
			if ws[ dic[z] + str(i) ].value == 'issue':
				ws.column_dimensions[dic[z]].width = len(str(ws[ dic[z] + str(i+1) ].value))
	for i in range(1,100):
		ws.column_dimensions[dic[0]].width = 5

	# for z in range(26):
	# 	ws.column_dimensions[dic[z]].width = 20
#	lagbe = "_A" if (choose == 6) else lagbe = ""
	workbook.save("Supervisor " + str(choose) + " queries" + ".xlsx")

#for i in range(1,10):
for i in [1,2,3,4,5,608,612,637,638,639,7,8,9]:
	TheGrandSorting(i)

# 	pass
	
#BLUE = FF00B0F0
#WHITE = 00000000
#YELLOW = FFFFFF00
#ORANGE = FFFFC000
#RED = FFC00000
